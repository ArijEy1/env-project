#!/usr/bin/env python3
"""
Convert the client's SEMS v0.4 question-bank workbook into normalized, committed
JSON seed files consumed by apps/api/src/database/import-sems-v04.ts.

The workbook itself is NOT committed (it lives outside the repo); this converter
is the reproducible bridge. Re-run it whenever the client sends a new version:

    python tools/convert-sems-xlsx.py \
        --src "C:/Users/Administrator/Downloads/SEMS_Draft_Question_Bank_v0.4.xlsx"

Output: apps/api/src/database/seed/sems-v04/*.json

Design notes
------------
* Domain is keyed off the Arabic `Domain` column via an explicit code map
  (DOMAIN_CODES). Prefixes are NOT reliable: NAT/SUP each span two domains.
  An unknown domain name is a hard error so a future version can't drift silently.
* Every question keeps its full 73-column row verbatim under `raw` for
  traceability; the mapped fields are a curated projection used by the engine.
* Answer options split on the Arabic semicolon (U+061B) or ASCII ';'.
* English text is NOT produced here — translations live in translations.json,
  authored/curated separately and merged at import time (en_review_status).
"""
import argparse
import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

# --- Stable domain code map (Arabic domain name -> short code + English) ------
# Faithful to the file: 17 distinct domains, including the near-duplicate
# nature/supply pairs the client separated (v0.3 vs v0.4 module additions).
DOMAIN_CODES = {
    "ملف الجهة والنطاق":            ("PROFILE",  "Entity Profile & Scope"),
    "أسئلة الانطباق والتفعيل":       ("APPLIC",   "Applicability & Activation"),
    "الحوكمة البيئية":              ("GOV",      "Environmental Governance"),
    "الالتزام التنظيمي":            ("REG",      "Regulatory Compliance"),
    "القياس وجودة البيانات":         ("DATA",     "Measurement & Data Quality"),
    "الطاقة والمناخ والانبعاثات":     ("CLIMATE",  "Energy, Climate & Emissions"),
    "المياه والصرف":                ("WATER",    "Water & Wastewater"),
    "النفايات والمواد الخطرة":        ("WASTE",    "Waste & Hazardous Materials"),
    "الطبيعة والمواقع الحساسة":       ("NATURE",   "Nature & Sensitive Sites"),
    "سلسلة الإمداد والموردون":        ("SUPPLY",   "Supply Chain & Suppliers"),
    "القياس والإفصاح":              ("DISCLOSE", "Measurement & Disclosure"),
    "التحسين والابتكار":            ("IMPROVE",  "Improvement & Innovation"),
    "النظام والتكامل":              ("SYSTEM",   "System & Integration"),
    "الهواء وجودة الأوساط":          ("AIR",      "Air & Ambient Quality"),
    "الطبيعة والموارد الحيوية":       ("BIORES",   "Nature & Biological Resources"),
    "سلسلة الإمداد والمشتريات":       ("PROCURE",  "Supply Chain & Procurement"),
    "التحسين والاقتصاد الدائري":      ("CIRCULAR", "Improvement & Circular Economy"),
}

# Which mapped column each question field reads from (source header -> field).
QUESTION_FIELDS = {
    "Question_ID": "id",
    "Status": "status",
    "Version": "version",
    "Question_Category": "category",
    "Question_Layer": "layer",
    "Measurement_Layer": "measurementLayer",
    "Domain": "_domainAr",
    "Sub_Domain": "subDomain",
    "Assessment_Element": "assessmentElement",
    "Question_Text_AR": "textAr",
    "Question_Purpose": "purposeAr",
    "Question_Typology": "typology",
    "Answer_Type": "answerType",
    "Answer_Options": "_answerOptionsRaw",
    "Applicability_Rule_ID": "applicabilityRuleId",
    "Applicability_Trigger": "applicabilityTrigger",
    "Applicability_Priority": "applicabilityPriority",
    "Environmental_Aspect": "environmentalAspect",
    "Impact_Category": "impactCategory",
    "Environmental_Impact": "environmentalImpact",
    "Risk_Link": "riskLink",
    "Evidence_Required": "evidenceRequired",
    "Min_Evidence_Level": "minEvidenceLevel",
    "Baseline_Required": "baselineRequired",
    "Trend_Required": "trendRequired",
    "Scoring_Treatment": "scoringTreatment",
    "Red_Flag_Logic": "redFlagLogic",
    "Interquestion_Check_ID": "interquestionCheckId",
    "Recommendation_ID": "recommendationId",
    "Guidance_Text": "guidanceAr",
    "Rule_Engine_ID": "ruleEngineId",
    "Decision_Term_Tags": "decisionTermTags",
    "Dependency_Parent_Questions": "dependencyParents",
    "Dependency_Child_Questions": "dependencyChildren",
    "Attribution_Required": "attributionRequired",
    "Attribution_Method": "attributionMethod",
    "Outcome_Threshold_ID": "outcomeThresholdId",
    "Maturity_Rubric_ID": "maturityRubricId",
    "Evidence_Freshness_Rule": "evidenceFreshnessRule",
    "Evidence_Conflict_Rule": "evidenceConflictRule",
    "Benchmark_Readiness_Criteria": "benchmarkReadinessCriteria",
    "Benchmarking_Method": "benchmarkingMethod",
    "Completion_Effort": "completionEffort",
    "Data_Availability_Level": "dataAvailability",
    "User_Difficulty": "userDifficulty",
    "Estimated_Completion_Time": "estimatedTime",
    "Basic_Guidance": "basicGuidanceAr",
    "Advanced_Guidance": "advancedGuidanceAr",
    "Retirement_Status": "retirementStatus",
    "KG_Node_ID": "kgNodeId",
}

# Reference sheets -> (output filename, id-ish first column for sanity).
REFERENCE_SHEETS = {
    "03_Rule_Engine_Spec": "rule-engine.json",
    "04_Decision_Dictionary": "decision-dictionary.json",
    "05_Conflict_Decision_Tree": "conflict-tree.json",
    "06_Maturity_Rubrics": "maturity-rubrics.json",
    "07_Attribution_Logic": "attribution.json",
    "08_Outcome_Thresholds": "outcome-thresholds.json",
    "09_Evidence_Freshness_Conflicts": "evidence-rules.json",
    "10_Benchmark_Readiness": "benchmark-readiness.json",
    "11_Dependency_Matrix": "dependencies.json",
    "12_Knowledge_Graph_Edges": "kg-edges.json",
    "18_Source_Lineage": "source-lineage.json",
}

SPLIT_RE = re.compile(r"[؛;]")


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s != "" else None


def split_options(raw):
    if not raw:
        return []
    parts = [p.strip() for p in SPLIT_RE.split(str(raw)) if p.strip()]
    out = []
    for p in parts:
        # Maturity options look like "0 لا يوجد" / "1 مبدئي"; pull the leading level.
        m = re.match(r"^(\d+)\s+(.*)$", p)
        if m:
            out.append({"value": m.group(1), "labelAr": m.group(2), "level": int(m.group(1))})
        else:
            out.append({"value": p, "labelAr": p, "level": None})
    return out


def header_row_index(rows):
    """First row (after the banner/subtitle) carrying >= 3 non-null cells."""
    for i, r in enumerate(rows):
        if i == 0:
            continue
        nonnull = sum(1 for c in r if c is not None and str(c).strip() != "")
        if nonnull >= 3:
            return i
    return 1


def read_sheet_records(ws):
    rows = list(ws.iter_rows(values_only=True))
    hi = header_row_index(rows)
    hdr = [clean(h) or f"col{j}" for j, h in enumerate(rows[hi])]
    records = []
    for r in rows[hi + 1:]:
        if not any(c is not None and str(c).strip() != "" for c in r):
            continue
        rec = {}
        for j, h in enumerate(hdr):
            val = clean(r[j]) if j < len(r) else None
            rec[h] = val
        records.append(rec)
    return hdr, records


def convert_questions(ws):
    rows = list(ws.iter_rows(values_only=True))
    hi = header_row_index(rows)
    hdr = [clean(h) or f"col{j}" for j, h in enumerate(rows[hi])]
    idx = {h: j for j, h in enumerate(hdr)}

    def cell(r, name):
        j = idx.get(name)
        return clean(r[j]) if j is not None and j < len(r) else None

    domains = {}  # code -> record
    questions = []
    order = 0
    for r in rows[hi + 1:]:
        if not any(c is not None and str(c).strip() != "" for c in r):
            continue
        q = {}
        for src, field in QUESTION_FIELDS.items():
            q[field] = cell(r, src)
        if not q.get("id"):
            continue

        dom_ar = q.pop("_domainAr")
        if dom_ar not in DOMAIN_CODES:
            sys.exit(f"ERROR: unknown Domain '{dom_ar}' for question {q['id']}. "
                     f"Add it to DOMAIN_CODES.")
        code, name_en = DOMAIN_CODES[dom_ar]
        q["domainId"] = code
        if code not in domains:
            domains[code] = {
                "id": code,
                "nameAr": dom_ar,
                "nameEn": name_en,
                "displayOrder": len(domains) + 1,
            }

        raw_opts = q.pop("_answerOptionsRaw")
        q["answerOptions"] = split_options(raw_opts)
        q["answerOptionsRaw"] = raw_opts
        q["displayOrder"] = order
        order += 1

        # Full verbatim row for traceability.
        q["raw"] = {h: (clean(r[j]) if j < len(r) else None) for h, j in idx.items()}
        questions.append(q)

    return list(domains.values()), questions


def main():
    ap = argparse.ArgumentParser()
    default_src = os.path.join(os.path.expanduser("~"), "Downloads",
                               "SEMS_Draft_Question_Bank_v0.4.xlsx")
    ap.add_argument("--src", default=default_src)
    here = os.path.dirname(os.path.abspath(__file__))
    default_out = os.path.join(here, "..", "apps", "api", "src", "database",
                               "seed", "sems-v04")
    ap.add_argument("--out", default=default_out)
    args = ap.parse_args()

    if not os.path.exists(args.src):
        sys.exit(f"Source workbook not found: {args.src}")
    outdir = os.path.abspath(args.out)
    os.makedirs(outdir, exist_ok=True)

    wb = openpyxl.load_workbook(args.src, data_only=True, read_only=True)

    def write(name, obj):
        path = os.path.join(outdir, name)
        with open(path, "w", encoding="utf-8") as fp:
            json.dump(obj, fp, ensure_ascii=False, indent=1)
        n = len(obj) if isinstance(obj, list) else "-"
        print(f"  wrote {name:26} ({n} records)")

    # Questions + domains
    ws_q = wb["01_Question_Bank_v0.4"]
    domains, questions = convert_questions(ws_q)
    write("domains.json", domains)
    write("questions.json", questions)

    # Reference sheets
    for sheet, fname in REFERENCE_SHEETS.items():
        if sheet not in wb.sheetnames:
            print(f"  WARN: sheet {sheet} missing")
            continue
        _hdr, recs = read_sheet_records(wb[sheet])
        write(fname, recs)

    meta = {
        "source": os.path.basename(args.src),
        "version": "v0.4",
        "questionCount": len(questions),
        "domainCount": len(domains),
    }
    write("_meta.json", meta)
    print(f"Done -> {outdir}")


if __name__ == "__main__":
    main()
